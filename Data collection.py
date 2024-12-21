#collection fo youtube comments 
from googleapiclient.discovery import build
import pandas as pd
import os
from openpyxl import load_workbook,Workbook

api_key = 'AIzaSyCabvbMPDb3cM0D3NepWzSE9Q_vx13uF_w'
video_id = 'mYPwMIbzFLI'

youtube = build('youtube', 'v3', developerKey=api_key)

def get_comments(video_id):
    request = youtube.commentThreads().list(
        part='snippet',
        videoId=video_id,
        maxResults=1000  # adjusting the number of comments
    )
    response = request.execute()

    comments = []
    for item in response['items']:
        comment = item['snippet']['topLevelComment']['snippet']['textDisplay']
        comments.append(comment)

    return comments
#loding in old or new workbook
comments = get_comments(video_id)
file_path = "youtube_comments_2021.xlsx"
if os.path.exists(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comments"
    ws['A1'] = "Comments"

next_row = ws.max_row + 1

for idx, comment in enumerate(comments, start=next_row):
    ws[f'A{idx}'] = comment
print(ws.max_row)

wb.save(file_path)